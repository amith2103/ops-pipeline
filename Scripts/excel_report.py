import pandas as pd 
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Load data from database
conn = sqlite3.connect("data/ops-pipeline.db")
df = pd.read_sql("SELECT * FROM service_requests", conn)
conn.close()

# key stats
total = len(df)
open_tickets = len(df[df["status"] == "Open"])
critical = len(df[df["priority"] =="Critical"])
resolved = len(df[df["status"] =="Resolved"])
by_region = df["region"].value_counts()
by_service = df["service_type"].value_counts()
by_priority = df["priority"].value_counts()

print("Data loaded successfully!")


# Create workbook 
wb = Workbook()

#----sheet1: summary----
ws1 = wb.active
ws1.tittle = "summary"

#Title 
ws1["A1"] = "Utility Service Request Dashboard"
ws1["A1"].font = Font(bold=True, size=16, color="FFFFFF")
ws1["A1"].fill = PatternFill("solid", fgColor="065A82")
ws1["A1"].alignment = Alignment(horizontal = "center")
ws1.merge_cells("A1:D1")

#Headers
headers = ["Metric", " Value"]
ws1.append([])
ws1.append(headers)
for col in range(1,3):
    ws1.cell(row=3, column=col).font = Font(bold=True, color = "FFFFFF")
    ws1.cell(row=3, column=col).fill = PatternFill("solid", start_color="1C7293")

# Stats
stats = [
    ["Total Requests", total],
    ["Open Tickets", open_tickets],
    ["Resolved Tickets", resolved],
    ["Critical priority", critical],
    ]

for row in stats:
    ws1.append(row)


# column widths
ws1.column_dimensions["A"].width = 25
ws1.column_dimensions["B"].width = 15    


#--Sheet 2: By Region ---
ws2 = wb.create_sheet("By Region")

ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws2["A1"].fill = PatternFill("solid", start_color="065A82")
ws2["A1"].alignment = Alignment(horizontal="center")
ws2.merge_cells("A1:B1")

ws2.append([])
ws2.append(["Region"," Count"])
for col in range(1,3):
    ws2.cell(row=3, column=col).font = Font(Font(bold=True, color="FFFFFF"))
    ws2.cell(row =3, column=col).fill = PatternFill("solid", start_color = "1C7293")

for region, count in by_region.items():
    ws2.append([region, count])

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 15


#  sheet 3: by service type----
ws3 = wb.create_sheet("By Service Type")

ws3["A1"] = "Requests by Service Type"
ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws3["A1"].fill = PatternFill("solid", start_color="065A82")
ws3["A1"].alignment = Alignment(horizontal="center")
ws3.merge_cells("A1:B1")

ws3.append([])
ws3.append(["Service Type", "Count"])
for col in range(1, 3):
    ws3.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
    ws3.cell(row=3, column=col).fill = PatternFill("solid", start_color="1C7293")

for service, count in by_service.items():
    ws3.append([service, count])

ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 15

# Save
wb.save("data/dashboard.xlsx")
print("Excel dashboard saved to data/dashboard.xlsx")



